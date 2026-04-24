import io
from datetime import datetime
from typing import List, Dict, Any

import pandas as pd
import streamlit as st
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

# =========================
# Excel書式
# =========================
blue_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
gray_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
green_fill = PatternFill(fill_type="solid", fgColor="C6E0B4")
red_fill = PatternFill(fill_type="solid", fgColor="F4B084")

center_align = Alignment(horizontal="center", vertical="center")

# =========================
# 詳細モード用：標準ボタン候補
# =========================
DEFAULT_STOP_WORK_OPTIONS = [
    "部品セット",
    "部品取り出し",
    "クランプ",
    "アンクランプ",
    "治具段取り",
    "工具交換",
    "測定準備",
    "清掃",
    "機械操作",
]

DEFAULT_RUNNING_WORK_OPTIONS = [
    "次部品準備",
    "外観検査",
    "バリ取り",
    "面取り",
    "測定",
    "洗浄",
    "梱包準備",
    "記録記入",
    "次工程準備",
]

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


st.set_page_config(page_title="マンマシンチャート簡単作成", layout="centered")


# =========================
# 初期化
# =========================
def init_session_state() -> None:
    defaults = {
        "events": [],
        "running": False,
        "finished": False,
        "start_ts": None,
        "current_state": "未開始",
        "current_detail_state": "",
        "work_name": "",
        "operator_name": "",
        "machine_name": "",
        "detail_mode": False,

        # 詳細モード：機械停止中の作業ボタン
        "custom_stop_input": "",
        "selected_stop_buttons": [
            "部品セット",
            "部品取り出し",
            "治具段取り",
        ],

        # 詳細モード：機械稼働中の作業ボタン
        "custom_running_input": "",
        "selected_running_buttons": [
            "次部品準備",
            "外観検査",
            "バリ取り",
        ],
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


init_session_state()


# =========================
# 共通関数
# =========================
def now_timestamp() -> datetime:
    return datetime.now()


def seconds_from_start(ts: datetime) -> float:
    if st.session_state.start_ts is None:
        return 0.0
    return max(0.0, (ts - st.session_state.start_ts).total_seconds())


def format_seconds(seconds: float) -> str:
    total = int(round(seconds))
    minutes = total // 60
    sec = total % 60
    return f"{minutes:02d}:{sec:02d}"


def human_label(state_code: str) -> str:
    mapping = {
        "work": "作業",
        "monitor": "監視・待ち",
        "end": "終了",
    }
    return mapping.get(state_code, state_code)


def detail_to_major(detail_state: str) -> str:
    if detail_state == "監視・待ち":
        return "monitor"
    return "work"


def detail_options() -> List[str]:
    return st.session_state.selected_stop_buttons


def running_detail_options() -> List[str]:
    return st.session_state.selected_running_buttons


def machine_label_from_human(state_code: str) -> str:
    # MVP前提：人が作業→機械停止、人が監視・待ち→機械稼働
    if state_code == "work":
        return "停止"
    if state_code == "monitor":
        return "稼働"
    return ""


def add_event(new_state: str, detail_state: str = "", machine_state: str = "") -> None:
    current_ts = now_timestamp()

    if st.session_state.start_ts is None:
        st.session_state.start_ts = current_ts

    elapsed = seconds_from_start(current_ts)

    if st.session_state.events:
        last_state = st.session_state.events[-1]["state"]
        last_detail = st.session_state.events[-1].get("detail_state", "")
        if last_state == new_state and last_detail == detail_state:
            return

    st.session_state.events.append(
        {
            "time": elapsed,
            "state": new_state,
            "detail_state": detail_state,
            "machine_state": machine_state,
            "timestamp": current_ts.strftime("%Y-%m-%d %H:%M:%S"),
        }
    )
    st.session_state.running = True
    st.session_state.finished = False
    st.session_state.current_state = human_label(new_state)
    st.session_state.current_detail_state = detail_state


def add_detail_event(detail_state: str, machine_state: str = "") -> None:
    major = detail_to_major(detail_state)

    if not machine_state:
        machine_state = machine_label_from_human(major)

    add_event(major, detail_state, machine_state)


def finish_measurement() -> None:
    if not st.session_state.events:
        return

    current_ts = now_timestamp()
    elapsed = seconds_from_start(current_ts)
    st.session_state.events.append(
        {
            "time": elapsed,
            "state": "end",
            "detail_state": "終了",
            "timestamp": current_ts.strftime("%Y-%m-%d %H:%M:%S"),
        }
    )
    st.session_state.running = False
    st.session_state.finished = True
    st.session_state.current_state = "計測終了"
    st.session_state.current_detail_state = ""


def undo_last() -> None:
    if not st.session_state.events:
        return

    st.session_state.events.pop()

    if not st.session_state.events:
        reset_all()
        return

    last_state = st.session_state.events[-1]["state"]
    last_detail = st.session_state.events[-1].get("detail_state", "")
    if last_state == "end":
        st.session_state.finished = True
        st.session_state.running = False
        st.session_state.current_state = "計測終了"
        st.session_state.current_detail_state = ""
    else:
        st.session_state.finished = False
        st.session_state.running = True
        st.session_state.current_state = human_label(last_state)
        st.session_state.current_detail_state = last_detail


def reset_all() -> None:
    st.session_state.events = []
    st.session_state.running = False
    st.session_state.finished = False
    st.session_state.start_ts = None
    st.session_state.current_state = "未開始"
    st.session_state.current_detail_state = ""


def build_intervals(events: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    intervals: List[Dict[str, Any]] = []
    if len(events) < 2:
        return intervals

    for i in range(len(events) - 1):
        start_event = events[i]
        end_event = events[i + 1]
        state = start_event["state"]
        if state == "end":
            continue

        start_sec = float(start_event["time"])
        end_sec = float(end_event["time"])
        duration = max(0.0, end_sec - start_sec)
        detail_state = start_event.get("detail_state", "") or human_label(state)

        intervals.append(
            {
                "No": i + 1,
                "区間開始(秒)": round(start_sec, 1),
                "区間終了(秒)": round(end_sec, 1),
                "区間時間(秒)": round(duration, 1),
                "状態(詳細)": detail_state,
                "人の状態": human_label(state),
                "機械の状態": start_event.get("machine_state", "") or machine_label_from_human(state),
                "メモ": "",
            }
        )
    return intervals


def build_log_dataframe(intervals: List[Dict[str, Any]]) -> pd.DataFrame:
    if not intervals:
        return pd.DataFrame(
            columns=[
                "No",
                "区間開始(秒)",
                "区間終了(秒)",
                "区間時間(秒)",
                "状態(詳細)",
                "人の状態",
                "機械の状態",
                "メモ",
            ]
        )
    return pd.DataFrame(intervals)


def build_summary_dataframe(intervals: List[Dict[str, Any]]) -> pd.DataFrame:
    if not intervals:
        return pd.DataFrame(columns=["指標", "値"])

    total_time = sum(x["区間時間(秒)"] for x in intervals)

    # 3分類
    stop_work_time = sum(
        x["区間時間(秒)"]
        for x in intervals
        if x["人の状態"] == "作業" and x["機械の状態"] == "停止"
    )

    parallel_work_time = sum(
        x["区間時間(秒)"]
        for x in intervals
        if x["人の状態"] == "作業" and x["機械の状態"] == "稼働"
    )

    monitor_time = sum(
        x["区間時間(秒)"]
        for x in intervals
        if x["人の状態"] == "監視・待ち"
    )

    machine_run_time = sum(
        x["区間時間(秒)"]
        for x in intervals
        if x["機械の状態"] == "稼働"
    )

    machine_stop_time = sum(
        x["区間時間(秒)"]
        for x in intervals
        if x["機械の状態"] == "停止"
    )

    monitor_segments = [
        x["区間時間(秒)"]
        for x in intervals
        if x["人の状態"] == "監視・待ち"
    ]

    def pct(value: float) -> float:
        return round((value / total_time * 100) if total_time else 0, 1)

    summary_rows = [
        {"指標": "総時間(秒)", "値": round(total_time, 1)},

        {"指標": "機械停止中の作業時間(秒)", "値": round(stop_work_time, 1)},
        {"指標": "機械停止中の作業比率(%)", "値": pct(stop_work_time)},

        {"指標": "機械稼働中の作業時間(秒)", "値": round(parallel_work_time, 1)},
        {"指標": "機械稼働中の作業比率(%)", "値": pct(parallel_work_time)},

        {"指標": "監視・待ち時間(秒)", "値": round(monitor_time, 1)},
        {"指標": "監視・待ち率(%)", "値": pct(monitor_time)},

        {"指標": "監視回数", "値": int(len(monitor_segments))},
        {"指標": "最長監視時間(秒)", "値": round(max(monitor_segments), 1) if monitor_segments else 0},

        {"指標": "機械稼働時間(秒)", "値": round(machine_run_time, 1)},
        {"指標": "機械稼働率(%)", "値": pct(machine_run_time)},

        {"指標": "機械停止時間(秒)", "値": round(machine_stop_time, 1)},
        {"指標": "機械停止率(%)", "値": pct(machine_stop_time)},
    ]

    return pd.DataFrame(summary_rows)


def build_raw_events_dataframe(events: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for i, event in enumerate(events, start=1):
        rows.append(
            {
                "No": i,
                "記録時刻": event["timestamp"],
                "経過秒": round(float(event["time"]), 1),
                "状態": human_label(event["state"]),
                "状態(詳細)": event.get("detail_state", "") or human_label(event["state"]),
            }
        )
    return pd.DataFrame(rows)


def create_excel_bytes(
    log_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    events_df: pd.DataFrame,
    work_name: str,
    operator_name: str,
    machine_name: str,
) -> bytes:
    wb = Workbook()

    # -------------------------
    # シート1: log
    # -------------------------
    ws_log = wb.active
    ws_log.title = "log"

    log_df = log_df.copy()
    if not log_df.empty:
        log_df = log_df.dropna(subset=["区間開始(秒)", "区間時間(秒)", "区間終了(秒)"])

    meta_rows = [
        ["作業名", work_name or ""],
        ["作業者名", operator_name or ""],
        ["設備名", machine_name or ""],
        ["出力日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [],
    ]
    for row in meta_rows:
        ws_log.append(row)

    header_row = 6
    for col_idx, col_name in enumerate(log_df.columns, start=1):
        cell = ws_log.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = center_align
        cell.border = thin_border

    data_start_row = header_row + 1
    for row_idx, (_, row) in enumerate(log_df.iterrows(), start=data_start_row):
        data = list(row)
        for col_idx, value in enumerate(data, start=1):
            cell = ws_log.cell(row=row_idx, column=col_idx, value=value)

            state = row["人の状態"]
            machine = row["機械の状態"]

            if state == "作業":
                cell.fill = blue_fill
            elif state == "監視・待ち":
                cell.fill = gray_fill

            if col_idx == 7:  # 機械の状態列
                if machine == "稼働":
                    cell.fill = green_fill
                elif machine == "停止":
                    cell.fill = red_fill

            cell.alignment = center_align
            cell.border = thin_border

    for row_num in range(1, ws_log.max_row + 1):
        ws_log.row_dimensions[row_num].height = 20

    for col_idx, col_name in enumerate(log_df.columns, start=1):
        max_length = len(str(col_name))
        for value in log_df[col_name]:
            max_length = max(max_length, len(str(value)))
        ws_log.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 5, 25)

    # -------------------------
    # グラフ用補助列
    # -------------------------
    chart_start_col = len(log_df.columns) + 3  # 2列空ける
    helper_headers = ["開始位置", "作業時間"]

    for idx, header in enumerate(helper_headers, start=chart_start_col):
        cell = ws_log.cell(row=header_row, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(log_df.iterrows(), start=data_start_row):
        start_val = row["区間開始(秒)"]
        duration_val = row["区間時間(秒)"]

        c1 = ws_log.cell(row=row_idx, column=chart_start_col, value=start_val)
        c2 = ws_log.cell(row=row_idx, column=chart_start_col + 1, value=duration_val)

        for c in [c1, c2]:
            c.alignment = center_align
            c.border = thin_border

    ws_log.column_dimensions[get_column_letter(chart_start_col)].width = 10
    ws_log.column_dimensions[get_column_letter(chart_start_col + 1)].width = 10

    # -------------------------
    # 95点版チャート
    # -------------------------
    if not log_df.empty:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "マンマシンチャート"
        # タイトルがグラフと重ならないように設定
        chart.title.overlay = False
        chart.x_axis.title = "区間No"
        chart.y_axis.title = "時間(秒)"
        chart.width = 25
        chart.height = max(6, len(log_df) * 0.5)
        chart.legend = None

        data = Reference(
            ws_log,
            min_col=chart_start_col,
            max_col=chart_start_col + 1,
            min_row=header_row,
            max_row=header_row + len(log_df),
        )

        cats = Reference(
            ws_log,
            min_col=1,   # No列
            min_row=header_row + 1,
            max_row=header_row + len(log_df),
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # 開始位置系列を透明化
        if len(chart.series) >= 1:
            chart.series[0].graphicalProperties.noFill = True
            chart.series[0].graphicalProperties.line.noFill = True

        # 作業時間系列を状態別に色分け
        if len(chart.series) >= 2:
            duration_series = chart.series[1]
            duration_series.dPt = []

            for i, (_, row) in enumerate(log_df.iterrows()):
                pt = DataPoint(idx=i)

                human_state = row["人の状態"]
                machine_state = row["機械の状態"]

                if human_state == "監視・待ち":
                    # 監視・待ち → グレー
                    pt.graphicalProperties.solidFill = "BFBFBF"

                elif human_state == "作業" and machine_state == "稼働":
                    # 並行作業 → 緑（ここが重要）
                    pt.graphicalProperties.solidFill = "70AD47"

                else:
                    # 作業＋機械停止 → 青
                    pt.graphicalProperties.solidFill = "5B9BD5"

                duration_series.dPt.append(pt)

        # No.1 を上にする
        #chart.x_axis.reverseOrder = True
        #chart.y_axis.reverseOrder = False

        # 棒を少し太くする
        chart.gapWidth = 50

        # --- 縦軸 (x_axis) の並び順を「表と同じ」にする設定 ---

        # 1. 並び順を「1が一番上」にする（前回 "maxMin" だったのを "minMax" に戻します）
        chart.x_axis.scaling.orientation = "maxMin"

        # 2. 時間軸（横軸）が上にいかないよう、一番下に固定する
        chart.x_axis.crosses = "min"

        # 横軸（時間軸）の基本設定
        chart.x_axis.scaling.min = 0

        # ===== ここが重要（順番＆軸）=====
        # 横棒グラフでは「時間軸」は value axis 扱い
        # openpyxlでは y_axis 側に効くケースがある

        # 1. 軸の線自体を表示させる（これが抜けていると目盛りが出ないことが多い）
        chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(w=9525))
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.numFmt = "0"

        # 横軸（数値軸）の設定
        chart.y_axis.delete = False
        chart.y_axis.axPos = "b"  # 軸をグラフの下側に配置
        chart.y_axis.numFmt = "General"

        # (重要) 目盛りの間隔を自動ではなく明示的に指定してみる（例: 10秒ごと）
        # これで数値が出るようになるケースが多いです
        # chart.y_axis.majorUnit = 10

        # --- 1. データから最大時間を自動取得 ---
        # 例：3列目（C列）に時間のデータがある場合
        # ws を ws_log に書き換えます
        # 数値（int または float）であるものだけを取り出すようにガードを入れます
        max_time = max(
            v for v in (ws_log.cell(row=i, column=3).value for i in range(2, ws_log.max_row + 1))
            if isinstance(v, (int, float))
        )

        # --- 2. 目盛りの間隔を計算 (前回のロジック) ---
        raw_interval = max_time / 10
        nice_units = [1, 2, 5, 10, 15, 20, 30, 60, 120, 300, 600, 1800, 3600]
        best_unit = next((u for u in nice_units if u >= raw_interval), nice_units[-1])

        # --- 3. グラフの設定に流し込む ---
        chart.y_axis.majorUnit = best_unit
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = math.ceil(max_time / best_unit) * best_unit

        # 念のため x_axis にも設定（補助）
        chart.x_axis.majorTickMark = "out"
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.delete = False

        # ===== ここまで =====

        # 配置（必ず最後）
        graph_anchor_col = chart_start_col + 4
        graph_anchor = f"{get_column_letter(graph_anchor_col)}6"
        ws_log.add_chart(chart, graph_anchor)

        # 凡例テキスト
        ws_log["J4"] = "凡例：青＝機械停止中の作業　緑＝機械稼働中の作業　グレー＝監視・待ち"

    # -------------------------
    # シート2: summary
    # -------------------------
    ws_summary = wb.create_sheet("summary")
    for col_idx, col_name in enumerate(summary_df.columns, start=1):
        cell = ws_summary.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=2):
        for col_idx, value in enumerate(list(row), start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border

    for col_idx, col_name in enumerate(summary_df.columns, start=1):
        max_length = len(str(col_name))
        for value in summary_df[col_name]:
            max_length = max(max_length, len(str(value)))
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 5, 25)

    # -------------------------
    # シート3: events
    # -------------------------
    ws_events = wb.create_sheet("events")
    for col_idx, col_name in enumerate(events_df.columns, start=1):
        cell = ws_events.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(events_df.iterrows(), start=2):
        for col_idx, value in enumerate(list(row), start=1):
            cell = ws_events.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border

    for col_idx, col_name in enumerate(events_df.columns, start=1):
        max_length = len(str(col_name))
        for value in events_df[col_name]:
            max_length = max(max_length, len(str(value)))
        ws_events.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 5, 28)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_chart_source(intervals: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for row in intervals:
        rows.append(
            {
                "開始": row["区間開始(秒)"],
                "終了": row["区間終了(秒)"],
                "長さ": row["区間時間(秒)"],
                "人の状態": row["人の状態"],
                "機械の状態": row["機械の状態"],
            }
        )
    return pd.DataFrame(rows)


# =========================
# UI
# =========================
st.title("マンマシンチャート簡単作成")
st.caption("MVP版：スマホで状態記録 → 画面確認 → Excel/CSV出力")

with st.expander("対象情報（あとでExcelに残ります）", expanded=True):
    st.session_state.work_name = st.text_input("作業名", value=st.session_state.work_name)
    st.session_state.operator_name = st.text_input("作業者名", value=st.session_state.operator_name)
    st.session_state.machine_name = st.text_input("設備名", value=st.session_state.machine_name)

with st.expander("記録モード設定", expanded=True):
    st.session_state.detail_mode = st.toggle(
        "詳細入力モード",
        value=st.session_state.detail_mode,
    )
    st.caption("標準モードは2ボタン、詳細モードは事前に選んだ作業ボタンを表示します。")

    st.markdown("##### 作業ボタン設定（機械停止中）")

    st.session_state.custom_stop_input = st.text_input(
        "候補にない作業名を追加（機械停止中・カンマ区切り）",
        value=st.session_state.custom_stop_input,
        placeholder="例：エアブロー,切粉除去,寸法確認",
    )

    custom_stop_items = [
        x.strip()
        for x in st.session_state.custom_stop_input.split(",")
        if x.strip()
    ]

    stop_candidates = list(dict.fromkeys(DEFAULT_STOP_WORK_OPTIONS + custom_stop_items))

    current_stop_default = list(
        dict.fromkeys(
            [x for x in st.session_state.selected_stop_buttons if x in stop_candidates]
            + custom_stop_items
        )
    )

    st.session_state.selected_stop_buttons = st.multiselect(
        "表示するボタン（機械停止中の作業）",
        options=stop_candidates,
        default=current_stop_default,
    )

    st.markdown("##### 作業ボタン設定（機械稼働中）")

    st.session_state.custom_running_input = st.text_input(
        "候補にない作業名を追加（機械稼働中・カンマ区切り）",
        value=st.session_state.custom_running_input,
        placeholder="例：次ロット準備,測定記録,外観確認",
    )

    custom_running_items = [
        x.strip()
        for x in st.session_state.custom_running_input.split(",")
        if x.strip()
    ]

    running_candidates = list(dict.fromkeys(DEFAULT_RUNNING_WORK_OPTIONS + custom_running_items))

    current_running_default = list(
        dict.fromkeys(
            [x for x in st.session_state.selected_running_buttons if x in running_candidates]
            + custom_running_items
        )
    )

    st.session_state.selected_running_buttons = st.multiselect(
        "表示するボタン（機械稼働中の作業）",
        options=running_candidates,
        default=current_running_default,
    )

st.markdown("---")

intervals = build_intervals(st.session_state.events)
log_df = build_log_dataframe(intervals)
summary_df = build_summary_dataframe(intervals)
raw_events_df = build_raw_events_dataframe(st.session_state.events)

elapsed_now = 0.0
if st.session_state.start_ts is not None and not st.session_state.finished:
    elapsed_now = seconds_from_start(now_timestamp())
elif st.session_state.events:
    elapsed_now = float(st.session_state.events[-1]["time"])

col_a, col_b = st.columns(2)
with col_a:
    current_state_text = st.session_state.current_state
    if st.session_state.current_detail_state:
        current_state_text = f"{st.session_state.current_state}（{st.session_state.current_detail_state}）"
    st.metric("現在の状態", current_state_text)
with col_b:
    st.metric("経過時間", format_seconds(elapsed_now))

st.markdown("### 記録ボタン")
if not st.session_state.detail_mode:
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        if st.button("作業開始", use_container_width=True):
            add_event("work", "作業")
            st.rerun()
    with btn_col2:
        if st.button("監視・待ち開始", use_container_width=True):
            add_event("monitor", "監視・待ち")
            st.rerun()
else:
    # 監視・待ちボタン
    # 人：監視・待ち ／ 機械：稼働
    st.markdown("#### 監視")
    if st.button("監視・待ち", use_container_width=True, key="monitor_detail"):
        add_detail_event("監視・待ち", "稼働")
        st.rerun()

    # 通常（機械停止中）の作業ボタン
    options_stop = detail_options()

    # 機械稼働中の作業ボタン
    options_run = running_detail_options()
    

    st.markdown("#### 作業（機械停止中）")

    cols_per_row = 2
    for start_idx in range(0, len(options_stop), cols_per_row):
        cols = st.columns(cols_per_row)
        for offset, detail_name in enumerate(options_stop[start_idx:start_idx + cols_per_row]):
            with cols[offset]:
                if st.button(
                    detail_name,
                    use_container_width=True,
                    key=f"stop_{detail_name}_{start_idx}_{offset}",
                ):
                    add_detail_event(detail_name, "停止")
                    st.rerun()

    # 機械稼働中ボタン（ある場合のみ表示）
    if options_run:
        st.markdown("#### 作業（機械稼働中）")

        for start_idx in range(0, len(options_run), cols_per_row):
            cols = st.columns(cols_per_row)
            for offset, detail_name in enumerate(options_run[start_idx:start_idx + cols_per_row]):
                with cols[offset]:
                    if st.button(
                        detail_name,
                        use_container_width=True,
                        key=f"run_{detail_name}_{start_idx}_{offset}",
                    ):
                        add_detail_event(detail_name, "稼働")
                        st.rerun()

btn_col3, btn_col4, btn_col5 = st.columns(3)
with btn_col3:
    if st.button("終了", use_container_width=True):
        finish_measurement()
        st.rerun()
with btn_col4:
    if st.button("1つ戻る", use_container_width=True):
        undo_last()
        st.rerun()
with btn_col5:
    if st.button("リセット", use_container_width=True):
        reset_all()
        st.rerun()

st.markdown("---")
st.subheader("記録履歴")
if raw_events_df.empty:
    st.info("まだ記録がありません。")
else:
    st.dataframe(raw_events_df, use_container_width=True, hide_index=True)

st.markdown("---")
st.subheader("集計結果")
if summary_df.empty:
    st.info("終了後に集計結果が表示されます。")
else:
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

st.markdown("---")
st.subheader("区間ログ")
if log_df.empty:
    st.info("区間ログはまだありません。")
else:
    st.dataframe(log_df, use_container_width=True, hide_index=True)

st.markdown("---")
st.subheader("簡易マンマシンチャート")
if not intervals:
    st.info("区間が確定すると、ここにチャートが表示されます。")
else:
    _ = build_chart_source(intervals)
    total_max = max(row["区間終了(秒)"] for row in intervals)
    scale = max(total_max, 1)
    width = 50

    def segment_bar(duration: float, state: str) -> str:
        length = max(1, int(round(duration / scale * width)))
        if state == "作業":
            return "🟦" * length
        return "⬜" * length

    def machine_bar(duration: float, state: str) -> str:
        length = max(1, int(round(duration / scale * width)))
        if state == "稼働":
            return "🟩" * length
        return "🟥" * length

    human_bar = ""
    machine_bar_all = ""
    for row in intervals:
        duration = float(row["区間時間(秒)"])
        human_bar += segment_bar(duration, row["人の状態"])
        machine_bar_all += machine_bar(duration, row["機械の状態"])

    st.markdown(f"**人**　{human_bar}")
    st.markdown(f"**機械** {machine_bar_all}")
    st.caption("凡例：人=🟦作業 / ⬜監視・待ち、機械=🟩稼働 / 🟥停止")

st.markdown("---")
st.subheader("出力")


def safe_name(text: str) -> str:
    cleaned = (text or "record").strip().replace(" ", "_")
    return cleaned or "record"


filename_base = f"{datetime.now().strftime('%Y-%m-%d_%H%M')}_{safe_name(st.session_state.work_name)}"

csv_bytes = log_df.to_csv(index=False).encode("utf-8-sig") if not log_df.empty else "".encode("utf-8-sig")
excel_bytes = create_excel_bytes(
    log_df=log_df,
    summary_df=summary_df,
    events_df=raw_events_df,
    work_name=st.session_state.work_name,
    operator_name=st.session_state.operator_name,
    machine_name=st.session_state.machine_name,
)

out_col1, out_col2 = st.columns(2)
with out_col1:
    st.download_button(
        label="Excel出力",
        data=excel_bytes,
        file_name=f"{filename_base}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        disabled=log_df.empty,
    )
with out_col2:
    st.download_button(
        label="CSV出力",
        data=csv_bytes,
        file_name=f"{filename_base}.csv",
        mime="text/csv",
        use_container_width=True,
        disabled=log_df.empty,
    )

st.caption("Excelには log / summary / events の3シートを出力します。logには詳細状態と大分類の両方を出力します。")
